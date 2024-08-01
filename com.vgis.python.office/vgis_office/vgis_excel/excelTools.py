"""
===================================
#!/usr/bin/python3.9
# -*- coding: utf-8 -*-
@Author: chenxw
@Email : gisfanmachel@gmail.com
@File: excelTools.py
@Date: Create in 2021/1/28 11:53
@Description: excel操作类
@ Software: PyCharm
===================================
"""
import os
import string

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side

from vgis_utils.vgis_file.fileTools import FileHelper
from vgis_utils.vgis_list.listTools import ListHelper

import win32com.client
from pathlib import Path

class ExcelHelper:

    @staticmethod
    # 需要安装了wps或office
    def xls2xlsx(input_filepath, output_filepath, keep_active=True):
        input_filepath = Path(input_filepath).resolve()
        output_filepath = Path(output_filepath).resolve()
        excel_app = win32com.client.Dispatch("Excel.Application")
        sheet = excel_app.Workbooks.Open(str(input_filepath))
        try:
            sheet.SaveAs(str(output_filepath), FileFormat=51)
            sheet.Close(0)
        except:
            sheet.Close(0)

        if not keep_active:
            excel_app.Quit()

    # 判断excel表里是否有某列
    @staticmethod
    def is_has_field_in_excel(excel_obj, field_name):
        is_find = False
        column_list = excel_obj.columns.values
        for field_index in range(len(column_list)):
            if column_list[field_index].strip() == field_name.strip():
                is_find = True
                break
        return is_find

    # 判断excel表里是否有某列(列名去中间空格）
    @staticmethod
    def is_has_field_of_remove_space_in_excel(excel_obj, field_name):
        is_find = False
        column_list = excel_obj.columns.values
        for field_index in range(len(column_list)):
            if column_list[field_index].strip().replace(" ", "") == field_name.strip().replace(" ", ""):
                is_find = True
                break
        return is_find

    @staticmethod
    # 通过excel字段名获取字段索引，若没找到，则为-1
    def get_field_index_by_name_in_excel(excel_obj, field_name):
        field_index_need = -1
        column_list = excel_obj.columns.values
        for field_index in range(len(column_list)):
            if column_list[field_index].strip() == field_name.strip():
                field_index_need = field_index
                break
        return field_index_need

    @staticmethod
    # 通过excel字段名（去掉中间空格）获取字段索引，若没找到，则为-1
    def get_field_index_by_name_of_remove_space_in_excel(excel_obj, field_name):
        field_index_need = -1
        column_list = excel_obj.columns.values
        for field_index in range(len(column_list)):
            if column_list[field_index].strip().replace(" ", "") == field_name.strip().replace(" ", ""):
                field_index_need = field_index
                break
        return field_index_need

    @staticmethod
    # 通过excel字段名（去掉中间空格和换行）获取字段索引，若没找到，则为-1
    def get_field_index_by_name_of_remove_space_and_breakline_in_excel(excel_obj, field_name):
        field_index_need = -1
        column_list = excel_obj.columns.values
        for field_index in range(len(column_list)):
            if column_list[field_index].strip().replace(" ", "").replace("\n", "") == field_name.strip().replace(" ",
                                                                                                                 "").replace(
                "\n", ""):
                field_index_need = field_index
                break
        return field_index_need

    @staticmethod
    # 通过excel字段名和行号获取字段值
    def get_field_value_by_name_in_excel(excel_obj, row_index, field_name):
        field_index_need = -1
        column_list = excel_obj.columns.values
        for field_index in range(len(column_list)):
            if column_list[field_index].strip() == field_name.strip():
                field_index_need = field_index
                break
        row_values = excel_obj.values[row_index]
        return row_values[field_index_need]

    @staticmethod
    # 获取excel的字段个数
    def get_field_count_in_excel(excel_obj):
        column_list = excel_obj.columns.values
        return len(column_list)

    @staticmethod
    # 为excel增加一列
    def add_field_in_excel(excel_obj, field_name):
        excel_obj[field_name] = None

    @staticmethod
    # 读取excel表内容指定行数据
    def read_excel_data_values_by_row(excel_obj, row_num):
        for row_index in range(len(excel_obj)):
            # excel内容从第二行开始
            if row_index == row_num - 1:
                result_row_values = excel_obj.values[row_index]
                return result_row_values

    @staticmethod
    #  读取excel表字段
    def read_excel_data_columns(excel_obj):
        result_row_values = excel_obj.columns
        return result_row_values

    @staticmethod
    # 构建新excel结果数据
    def build_data_excel(result_excel_path, all_data_list):
        file_dir_exists = os.path.exists(os.path.dirname(result_excel_path))
        if file_dir_exists is False:
            os.makedirs(os.path.dirname(result_excel_path))
        if os.path.exists(result_excel_path):
            os.remove(result_excel_path)
        if len(all_data_list) > 0:
            with pd.ExcelWriter(result_excel_path) as writer:
                for i in range(len(all_data_list)):
                    each_data_list = all_data_list[i]
                    if len(each_data_list) > 0:
                        df = pd.DataFrame(each_data_list,
                                          columns=ListHelper.get_key_name_str(each_data_list[0]))
                        sheet_name_str = "sheet" + str(i + 1)
                        df.to_excel(writer, sheet_name=sheet_name_str, index=False, startrow=0, encoding="utf_8_sig")
                # writer.save()
                # writer.close()
        else:
            # 创建空白excel
            fd = open(result_excel_path, 'w')
            fd.close()

    # 更新excel数据
    # 模板文件路径excel_template_path
    # 新文件路径 excel_target_path
    # 数据所在的sheet：sheet_name
    # 更新数据data_list:[{},{},{}]
    # 更新值开始行 value_start_row
    # 更新值开始列 value_start_col
    # 设置border结束列，border_end_col 这个大于等于更新值的结束列，因为excel可能有些其他列存在
    @staticmethod
    def update_excel(excel_template_path, excel_target_path, sheet_name, data_list, value_start_row, value_start_col,
                     border_end_col):
        FileHelper.copy_file(excel_template_path, excel_target_path)
        wb = openpyxl.load_workbook(excel_target_path)  # 打开excel工作簿
        # ws = wb.active  # 获取活跃sheet
        # ws = wb[wb.sheetnames[0]] #获取第一个sheet
        ws = wb[sheet_name]  # 获取指定名称的sheet
        # ws.cell(2, 1).value = 4
        # ws.cell(2, 2).value = 5
        # ws.cell(2, 3).value = 6
        # ws.cell(3, 1).value = 7
        # ws.cell(3, 2).value = 8
        # ws.cell(3, 3).value = 9
        data_row_index = 0
        for each_data in data_list:
            data_col_index = 0
            for key, value in each_data.items():
                ws.cell(value_start_row + data_row_index, value_start_col + data_col_index).value = value
                data_col_index += 1
            data_row_index += 1

        # border style
        border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
        # cell format , horizontal: 水平对齐；vertical：中心对齐；wrap_text：自动换行、
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        value_end_row = value_start_row + len(data_list) - 1
        # set cell border
        for row in range(value_start_row, value_end_row + 1):
            for col in range(value_start_col, border_end_col + 1):
                ws.cell(row, col).border = border
                ws.cell(row, col).alignment = align

        wb.save(excel_target_path)  # 保存

    @staticmethod
    # 根据Excel列索引值生成列字母名
    def get_column_name(column_index):
        ret = ''
        ci = column_index - 1
        index = ci // 26
        if index > 0:
            ret += ExcelHelper.get_column_name(index)
        ret += string.ascii_uppercase[ci % 26]
        return ret

    @staticmethod
    def get_field_name_by_field_index(excel_obj, field_index):
        field_name_need = ""
        column_list = excel_obj.columns.values
        for tmp_index in range(len(column_list)):
            if field_index == tmp_index:
                field_name_need = column_list[field_index].strip().replace(" ", "")
                break
        return field_name_need


if __name__ == '__main__':
    data_list = []
    obj = {}
    obj["序号"] = 1
    obj['批改标志'] = "增加"
    obj["分组组别"] = "2组"
    obj['被保险人名称'] = "中核1"
    obj['证件类型'] = '居民身份证'
    obj['证件号码'] = "23325252352"
    obj['性别'] = "男"
    obj['年龄'] = 25
    obj['出生日期\n(YYYY-MM-DD)'] = "2034-45-23"
    obj['职业代码人员类别'] = "100023"
    data_list.append(obj)
    obj = {}
    obj["序号"] = 2
    obj['批改标志'] = "减少"
    obj["分组组别"] = "3组"
    obj['被保险人名称'] = "中核3"
    obj['证件类型'] = '居民身份证'
    obj['证件号码'] = "54544662525235332"
    obj['性别'] = "女"
    obj['年龄'] = 32
    obj['出生日期\n(YYYY-MM-DD)'] = "2294-45-23"
    obj['职业代码人员类别'] = "100029"
    data_list.append(obj)

    excel_template_path = "E:\\系统开发\\中核保险辅助出单\\zwbx_fzcd_service\\my_project\\my_app\\zhbxfzcd_static\\template\\POLICY_PEOPLE_INFORMATION_TEMPLATE.xlsx"
    excel_target_path = "E:\\系统开发\\中核保险辅助出单\\zwbx_fzcd_service\\my_project\\my_app\\zhbxfzcd_static\\template\\POLICY_PEOPLE_INFORMATION_NEW8.xlsx"
    ExcelHelper.update_excel(excel_template_path, excel_target_path, "人员清单", data_list, 2, 1, 10)

    excel_template_path = "E:\\系统开发\\中核保险辅助出单\\zwbx_fzcd_service\\my_project\\my_app\\zhbxfzcd_static\\template\\ENDORSEMENT_PEOPLE_INFORMATION_TEMPLATE.xlsx"
    excel_target_path = "E:\\系统开发\\中核保险辅助出单\\zwbx_fzcd_service\\my_project\\my_app\\zhbxfzcd_static\\template\\ENDORSEMENT_PEOPLE_INFORMATION_NEW7.xlsx"
    ExcelHelper.update_excel(excel_template_path, excel_target_path, "人员清单", data_list, 5, 1, 83)
