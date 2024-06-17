#!/usr/bin/python3.9
# -*- coding: utf-8 -*-
# @Time    :  2022/9/28 10:36
# @Author  : chenxw
# @Email   : gisfanmachel@gmail.com
# @File    : worddir2xls.py
# @Descr   : 
# @Software: PyCharm

from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, borders, colors, Side

# 原始word文件和要生成的Excel文件
fn_word = 'E:\\tt.docx'
fn_excel = fn_word[:-5] + '2.xlsx'

# 创建空白Excel文件，获取第一个空白工作表
wb = Workbook()
ws = wb.worksheets[0]
# 添加表头
ws.append(['一级目录', '二级目录', '三级目录', '四级目录', '五级目录', '六级目录'])

first2, first3, first4, first5, first6 = 1, 1, 1, 1, 1
# 遍历word文件中的所有段落文本
for p in Document(fn_word).paragraphs:
    # 删除段落文本两侧的空白字符
    txt = p.text.rstrip()
    # 直接跳过空行
    if not txt:
        continue

    # 添加为一级标题
    if p.style.name == 'Heading 1':
        first2 = 1
        print(p.text)
        ws.append([txt, '', '', '', '', ''])
    # 添加为二级标题
    elif p.style.name == 'Heading 2':
        print(p.text)
        if first2 == 1:
            # 第一个二级目录，不添加新行
            # 直接和一级目录写在同一行
            # print(vgis_txt)
            list(ws.rows)[-1][1].value = txt
            first2 = first2 + 1
        elif first2 > 1:
            # 新增一行添加为二级标题
            # print(vgis_txt)
            ws.append(['', txt, '', '', '', ''])
        first3 = 1
    # 添加为三级标题
    elif p.style.name == 'Heading 3':
        print(p.text)
        if first3 == 1:
            # 第一个三级目录，不添加新行，直接写和二级在同一行
            # print(vgis_txt)
            list(ws.rows)[-1][2].value = txt
            first3 = first3 + 1
        elif first3 > 1:
            # 新增一行添加为三级目录
            print(txt)
            ws.append(['', '', txt, '', '', ''])
    # 添加为四级标题
    elif p.style.name == 'Heading 4':
        print(p.text)
        if first4 == 1:
            # 第一个四级目录，不添加新行，直接写和三级在同一行
            # print(vgis_txt)
            list(ws.rows)[-1][3].value = txt
            first4 = first4 + 1
        elif first4 > 1:
            # 新增一行添加为四级目录
            print(txt)
            ws.append(['', '', '', txt, '', ''])
    # 添加为五级标题
    elif p.style.name == 'Heading 5':
        print(p.text)
        if first5 == 1:
            # 第一个五级目录，不添加新行，直接写和四级在同一行
            # print(vgis_txt)
            list(ws.rows)[-1][4].value = txt
            first5 = first5 + 1
        elif first5 > 1:
            # 新增一行添加为五级目录
            print(txt)
            ws.append(['', '', '', '', txt, ''])
    # 添加为六级标题
    elif p.style.name == 'Heading 6':
        print(p.text)
        if first6 == 1:
            # 第一个六级目录，不添加新行，直接写和五级在同一行
            # print(vgis_txt)
            list(ws.rows)[-1][5].value = txt
            first6 = first6 + 1
        elif first6 > 1:
            # 新增一行添加为六级目录
            print(txt)
            ws.append(['', '', '', '', '', txt])

wb.save(fn_excel)

# 设置单元格边框
side = Side(border_style=borders.BORDER_THICK,
            color=colors.BLACK)
for row in ws.rows:
    for cell in row:
        cell.border = Border(top=side, bottom=side,
                             left=side, right=side)
start1, start2 = 1, 1
for index, row in enumerate(ws.rows, start=1):
    # 合并第一列的指定单元格，水平左对齐，垂直居中对齐
    if row[0].value:
        if index > start1:
            ws.merge_cells(f'A{start1}:A{index - 1}')
            ws[f'A{start1}'].alignment = Alignment(horizontal='left',
                                                   vertical='center')
        start1 = index
    # 合并第二列的指定单元格，水平左对齐，垂直居中对齐
    if row[1].value:
        if index > start2:
            ws.merge_cells(f'B{start2}:B{index - 1}')
            ws[f'B{start2}'].alignment = Alignment(horizontal='left',
                                                   vertical='center')
        start2 = index
if index > start1:
    ws.merge_cells(f'A{start1}:A{index}')
    ws[f'A{start1}'].alignment = Alignment(horizontal='left',
                                           vertical='center')

if index > start2:
    ws.merge_cells(f'B{start2}:B{index}')
    ws[f'B{start2}'].alignment = Alignment(horizontal='left',
                                           vertical='center')

wb.save(fn_excel)
